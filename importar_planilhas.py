import json
import os
from datetime import datetime
import openpyxl

ARQUIVO_OS = "dados_os.json"
ARQUIVO_TROCA = "dados_troca.json"
ARQUIVO_OBS = "dados_observacoes.json"
CAMINHO_PLANILHAS = r"D:\planilhas do adao"


def importar_os():
    caminho = os.path.join(CAMINHO_PLANILHAS, "planilha Chamados de Os(OS).xlsx")
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(min_row=2, values_only=True))
    dados = []
    for row in linhas:
        if not row[0]:
            continue
        dados.append({
            "numero_os": str(row[0] or ""),
            "serial": str(row[1] or ""),
            "modelo": str(row[2] or ""),
            "cep": str(row[3] or ""),
            "localidade": str(row[4] or ""),
            "abertura": str(row[5] or ""),
            "previsao": str(row[6] or ""),
            "finalizado": str(row[7] or ""),
            "descricao": str(row[8] or ""),
            "observacao": str(row[9] or ""),
            "status": str(row[10] or ""),
            "ocorrencia": str(row[11] or ""),
        })

    with open(ARQUIVO_OS, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    print(f"OS importadas: {len(dados)}")
    return dados


def importar_troca():
    caminho = os.path.join(CAMINHO_PLANILHAS, "Planilha troca antecipada.xlsx")
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(min_row=2, values_only=True))
    dados = []
    for row in linhas:
        if not row[4]:
            continue
        dados.append({
            "loja": str(row[0] or ""),
            "cep": str(row[1] or ""),
            "uf": str(row[2] or ""),
            "modelo": str(row[3] or ""),
            "serial": str(row[4] or ""),
            "data_leitura": str(row[5] or ""),
            "percentual_toner": str(row[6] or ""),
            "cor_toner": str(row[7] or ""),
            "ip": str(row[8] or ""),
            "data_atendimento": str(row[9] or ""),
            "descricao": str(row[10] or ""),
            "status": str(row[11] or ""),
            "data_finalizacao": str(row[12] or ""),
            "email": str(row[13] or ""),
            "contato": str(row[14] or ""),
            "retorno_acao": str(row[15] or ""),
            "retorno": str(row[16] or ""),
            "col18": str(row[17] or ""),
        })

    with open(ARQUIVO_TROCA, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)
    print(f"Trocas importadas: {len(dados)}")
    return dados


def carregar_os():
    if os.path.exists(ARQUIVO_OS):
        with open(ARQUIVO_OS, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def carregar_troca():
    if os.path.exists(ARQUIVO_TROCA):
        with open(ARQUIVO_TROCA, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def obter_resumo_os():
    dados = carregar_os()
    total = len(dados)
    finalizadas = sum(1 for d in dados if d.get("status", "").lower() == "finalizada")
    abertas = total - finalizadas
    modelos = {}
    for d in dados:
        m = d.get("modelo", "N/A")
        modelos[m] = modelos.get(m, 0) + 1
    modelos_ord = sorted(modelos.items(), key=lambda x: -x[1])[:10]

    return {
        "total": total,
        "finalizadas": finalizadas,
        "abertas": abertas,
        "modelos": [{"nome": k, "total": v} for k, v in modelos_ord],
    }


def obter_resumo_troca():
    dados = carregar_troca()
    total = len(dados)
    por_uf = {}
    por_cor = {}
    por_modelo = {}
    for d in dados:
        uf = d.get("uf", "N/A")
        por_uf[uf] = por_uf.get(uf, 0) + 1
        cor = d.get("cor_toner", "N/A")
        por_cor[cor] = por_cor.get(cor, 0) + 1
        m = d.get("modelo", "N/A")
        por_modelo[m] = por_modelo.get(m, 0) + 1

    toner_baixo = sum(1 for d in dados if d.get("percentual_toner", "0").isdigit() and int(d["percentual_toner"]) <= 20)

    return {
        "total": total,
        "toner_baixo": toner_baixo,
        "por_uf": [{"uf": k, "total": v} for k, v in sorted(por_uf.items(), key=lambda x: -x[1])],
        "por_cor": [{"cor": k, "total": v} for k, v in sorted(por_cor.items(), key=lambda x: -x[1])],
        "por_modelo": [{"nome": k, "total": v} for k, v in sorted(por_modelo.items(), key=lambda x: -x[1])[:10]],
    }


if __name__ == "__main__":
    importar_os()
    importar_troca()
    print("Importacao concluida!")


def carregar_observacoes():
    if os.path.exists(ARQUIVO_OBS):
        with open(ARQUIVO_OBS, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def salvar_observacoes(dados):
    with open(ARQUIVO_OBS, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def adicionar_observacao(serial, texto):
    obs = carregar_observacoes()
    if serial not in obs:
        obs[serial] = []
    obs[serial].append({
        "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "texto": texto
    })
    salvar_observacoes(obs)


def obter_observacoes(serial):
    obs = carregar_observacoes()
    return obs.get(serial, [])
