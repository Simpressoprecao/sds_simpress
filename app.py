import json
import os
import tempfile
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from sds_scraper import SDSScraper, salvar_json, carregar_json
from relatorios import gerar_excel, carregar_config, salvar_config, enviar_email, MENSAGEM_PADRAO
from database import Database, carregar_config_db, salvar_config_db
from importar_planilhas import carregar_os, carregar_troca, obter_resumo_os, obter_resumo_troca, adicionar_observacao, obter_observacoes, carregar_observacoes

app = Flask(__name__)
ARQUIVO_DADOS = "dados_dispositivo.json"
USUARIO = "ardlopes@simpress.com.br"
SENHA = "Rodrigo2021."
_database = None


def obter_db():
    global _database
    if _database is None:
        try:
            _database = Database()
            _database.criar_tabelas()
            _database.fechar()
            _database.conexao = None
        except Exception:
            _database = None
    return _database


def salvar_no_banco(dados):
    db = obter_db()
    if db:
        try:
            db.conectar()
            db.salvar_dispositivo(dados)
            db.fechar()
            return True
        except Exception:
            db = None
    return False


def listar_dispositivos_db():
    db = obter_db()
    if db:
        try:
            db.conectar()
            dados = db.listar_dispositivos()
            db.fechar()
            return dados
        except Exception:
            pass
    return None


def obter_scraper():
    return SDSScraper(USUARIO, SENHA, headless=True)


@app.route("/")
def index():
    dados = []
    resumo = {"total": 0, "pb_total": 0, "color_total": 0, "modelos": {}, "toner_baixo": 0, "modelos_ord": [], "alertas_serial": [], "alertas_qtd": []}

    db_dados = listar_dispositivos_db()
    if db_dados:
        for d in db_dados:
            modelo = d.get("modelo", "N/A")
            resumo["modelos"][modelo] = resumo["modelos"].get(modelo, 0) + 1
            pb = str(d.get("contador_pb", "0"))
            color = str(d.get("contador_color", "0"))
            try:
                resumo["pb_total"] += int(pb) if pb and pb.isdigit() else 0
                resumo["color_total"] += int(color) if color and color.isdigit() else 0
            except:
                pass
            consum = d.get("consumiveis", {}).get("consumiveis", [])
            for c in consum:
                nv = c.get("nivel", "0%").replace("%", "")
                if nv.isdigit() and int(nv) <= 20:
                    resumo["toner_baixo"] += 1
                    break
            qtd_alertas = int(d.get("total_alertas", 0))
            if qtd_alertas > 0:
                resumo["alertas_serial"].append(d.get("serial", "?")[:12])
                resumo["alertas_qtd"].append(qtd_alertas)
            resumo["total"] += 1
        dados = db_dados
    elif os.path.exists(ARQUIVO_DADOS):
        raw = carregar_json(ARQUIVO_DADOS)
        if not isinstance(raw, list):
            raw = [raw]
        for d in raw:
            disp = d.get("dispositivo", {})
            contagens_raw = d.get("contagens", {})
            contagens = (contagens_raw.get("contagens", {})
                         if isinstance(contagens_raw, dict) else {})
            if not contagens:
                contagens = disp.get("contagens", {})
            consum = d.get("consumiveis", {}).get("consumiveis", [])
            alertas = d.get("alertas", {}).get("frequencia_alertas", [])
            modelo = disp.get("modelo", "N/A")
            resumo["modelos"][modelo] = resumo["modelos"].get(modelo, 0) + 1
            pb = contagens.get("P\u00e1ginas monocrom\u00e1ticas", "0").replace(".", "")
            color = contagens.get("Colorido (equivalente A4)", "0").replace(".", "")
            try:
                resumo["pb_total"] += int(pb) if pb and pb.isdigit() else 0
                resumo["color_total"] += int(color) if color and color.isdigit() else 0
            except: pass
            for c in consum:
                nv = c.get("nivel", "0%").replace("%", "")
                if nv.isdigit() and int(nv) <= 20:
                    resumo["toner_baixo"] += 1
                    break
            total_alertas = 0
            for a in alertas:
                val = a.get("contagem", 0)
                if isinstance(val, str):
                    val = val.strip()
                    if val.isdigit():
                        total_alertas += int(val)
                elif isinstance(val, (int, float)):
                    total_alertas += int(val)
            if total_alertas > 0:
                resumo["alertas_serial"].append(disp.get("N\u00famero de s\u00e9rie", "?")[:12])
                resumo["alertas_qtd"].append(total_alertas)
            resumo["total"] += 1
            dados.append(d)

    resumo["modelos_ord"] = sorted(resumo["modelos"].items(), key=lambda x: -x[1])

    return render_template("index.html", dispositivos=dados, resumo=resumo)


@app.route("/buscar", methods=["POST"])
def buscar():
    serial = request.form.get("serial", "").strip()
    if not serial:
        return jsonify({"erro": "Informe um numero de serie ou ID"}), 400

    scraper = obter_scraper()
    try:
        scraper.login()
        if serial.isdigit():
            dados = scraper.extrair_tudo(serial)
        else:
            dados = scraper.processar_serial(serial)
        if dados:
            salvar_no_banco(dados)
            if os.path.exists(ARQUIVO_DADOS):
                existentes = carregar_json(ARQUIVO_DADOS)
                if not isinstance(existentes, list):
                    existentes = [existentes]
                existentes.append(dados)
                salvar_json(existentes, ARQUIVO_DADOS)
            else:
                salvar_json([dados], ARQUIVO_DADOS)
            return jsonify(dados)
        return jsonify({"erro": "Dispositivo nao encontrado ou modelo fora do padrao (E52645dn/E877dn)"}), 404
    finally:
        scraper._fechar()


@app.route("/api/dispositivos")
def listar_dispositivos():
    db_dados = listar_dispositivos_db()
    if db_dados:
        return jsonify(db_dados)

    if os.path.exists(ARQUIVO_DADOS):
        dados = carregar_json(ARQUIVO_DADOS)
        if not isinstance(dados, list):
            dados = [dados]
        resumo = []
        for d in dados:
            disp = d.get("dispositivo", {})
            consum = d.get("consumiveis", {}).get("consumiveis", [])
            toner_baixo = [
                c for c in consum
                if c.get("nivel", "0%").replace("%", "").isdigit()
                and int(c.get("nivel", "0%").replace("%", "")) <= 20
            ]
            resumo.append({
                "id": disp.get("id"),
                "serial": disp.get("N\u00famero de s\u00e9rie", ""),
                "modelo": disp.get("modelo", ""),
                "localizacao": disp.get("Localiza\u00e7\u00e3o", ""),
                "zona": disp.get("Zona", ""),
                "cliente": disp.get("breadcrumbs", [{}])[0].get("nome", "") if disp.get("breadcrumbs") else "",
                "ultima_atualizacao": disp.get("\u00daltima atualiza\u00e7\u00e3o", ""),
                "toner_baixo": toner_baixo,
                "ip": disp.get("Endere\u00e7o IP", ""),
            })
        return jsonify(resumo)
    return jsonify([])


@app.route("/api/dispositivo/<serial>")
def dispositivo_por_serial(serial):
    db = obter_db()
    if db:
        try:
            db.conectar()
            dados = db.buscar_dispositivo(serial)
            db.fechar()
            if dados:
                return jsonify(dados)
        except Exception:
            pass

    if os.path.exists(ARQUIVO_DADOS):
        dados = carregar_json(ARQUIVO_DADOS)
        if not isinstance(dados, list):
            dados = [dados]
        for d in dados:
            disp = d.get("dispositivo", {})
            if disp.get("N\u00famero de s\u00e9rie", "") == serial or disp.get("id") == serial:
                return jsonify(d)
    return jsonify({"erro": "Nao encontrado"}), 404


@app.route("/exportar-excel")
def exportar_excel():
    db = obter_db()
    if db:
        try:
            db.conectar()
            linhas = db.obter_dados_para_excel()
            db.fechar()
            if linhas:
                dados_excel = []
                for row in linhas:
                    serial = row["serial"]
                    ult_atual = row["ultima_atualizacao"] or ""
                    mes = ano = ""
                    for fmt in ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"]:
                        try:
                            dt = datetime.strptime(ult_atual, fmt)
                            mes = str(dt.month)
                            ano = str(dt.year)
                            break
                        except ValueError:
                            continue
                    dados_excel.append({
                        "dispositivo": {
                            "N\u00famero de s\u00e9rie": serial,
                            "\u00daltima atualiza\u00e7\u00e3o": ult_atual,
                            "contagens": {
                                "P\u00e1ginas monocrom\u00e1ticas": row.get("contador_pb", ""),
                                "Colorido (equivalente A4)": row.get("contador_color", ""),
                            }
                        }
                    })
                caminho = gerar_excel(dados_excel)
                return send_file(caminho, as_attachment=True, download_name=os.path.basename(caminho))
        except Exception as e:
            print(f"Erro MySQL no exportar-excel: {e}")

    if os.path.exists(ARQUIVO_DADOS):
        dados = carregar_json(ARQUIVO_DADOS)
        if not isinstance(dados, list):
            dados = [dados]
        caminho = gerar_excel(dados)
        return send_file(caminho, as_attachment=True, download_name=os.path.basename(caminho))
    return jsonify({"erro": "Nenhum dado para exportar"}), 404


@app.route("/config-email", methods=["GET", "POST"])
def config_email():
    if request.method == "POST":
        config = {
            "smtp_host": request.form.get("smtp_host", ""),
            "smtp_port": int(request.form.get("smtp_port", 587)),
            "email_from": request.form.get("email_from", ""),
            "email_pass": request.form.get("email_pass", ""),
            "email_to": request.form.get("email_to", ""),
            "mensagem": request.form.get("mensagem", MENSAGEM_PADRAO),
        }
        salvar_config(config)
        return jsonify({"ok": True, "mensagem": "Configuracao salva!"})
    config = carregar_config()
    config.setdefault("mensagem", MENSAGEM_PADRAO)
    return render_template("config_email.html", config=config)


@app.route("/config-db", methods=["GET", "POST"])
def config_db():
    if request.method == "POST":
        config = {
            "host": request.form.get("host", "localhost"),
            "port": int(request.form.get("port", 3306)),
            "user": request.form.get("user", "root"),
            "password": request.form.get("password", ""),
            "database": request.form.get("database", "sds_simpress"),
        }
        salvar_config_db(config)
        global _database
        _database = None
        try:
            db = Database(config)
            db.criar_tabelas()
            db.fechar()
            return jsonify({"ok": True, "mensagem": "Conectado ao MySQL com sucesso!"})
        except Exception as e:
            return jsonify({"ok": True, "mensagem": f"Configuracao salva, mas conexao falhou: {e}"})
    config = carregar_config_db()
    return render_template("config_db.html", config=config)


@app.route("/schema.sql")
def baixar_schema():
    return send_file("schema.sql", as_attachment=True, download_name="schema.sql")


@app.route("/enviar-relatorio", methods=["POST"])
def enviar_relatorio():
    para = request.form.get("para", "").strip()
    if not para:
        config = carregar_config()
        para = config.get("email_to", "")

    if not para:
        return jsonify({"erro": "Destinatario nao configurado"}), 400

    conf = carregar_config()
    if not conf.get("smtp_host") or not conf.get("email_from") or not conf.get("email_pass"):
        return jsonify({"erro": "Configure os dados SMTP em Configurar Email primeiro"}), 400

    dados = None
    db = obter_db()
    if db:
        try:
            db.conectar()
            linhas = db.obter_dados_para_excel()
            if linhas:
                dados = []
                for row in linhas:
                    serial = row["serial"]
                    ult_atual = row["ultima_atualizacao"] or ""
                    mes = ano = ""
                    for fmt in ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"]:
                        try:
                            dt = datetime.strptime(ult_atual, fmt)
                            mes = str(dt.month)
                            ano = str(dt.year)
                            break
                        except ValueError:
                            continue
                    dados.append({
                        "dispositivo": {
                            "N\u00famero de s\u00e9rie": serial,
                            "\u00daltima atualiza\u00e7\u00e3o": ult_atual,
                            "contagens": {
                                "P\u00e1ginas monocrom\u00e1ticas": row.get("contador_pb", ""),
                                "Colorido (equivalente A4)": row.get("contador_color", ""),
                            }
                        }
                    })
            db.fechar()
        except Exception:
            pass

    if not dados and os.path.exists(ARQUIVO_DADOS):
        dados_json = carregar_json(ARQUIVO_DADOS)
        if not isinstance(dados_json, list):
            dados_json = [dados_json]
        dados = dados_json

    if not dados:
        return jsonify({"erro": "Nenhum dado disponivel"}), 404

    data_str = datetime.now().strftime("%d%m%Y_%H%M%S")
    caminho_excel = f"relatorio_contadores_{data_str}.xlsx"
    gerar_excel(dados, caminho_excel)

    try:
        ok = enviar_email(para, "Relatorio de Contadores - Dispositivos",
                          conf.get("mensagem", MENSAGEM_PADRAO), caminho_excel, conf)
        if ok:
            return jsonify({"ok": True, "mensagem": f"Relatorio enviado para {para}!"})
        return jsonify({"erro": "Falha ao enviar email"}), 500
    except Exception as e:
        return jsonify({"erro": f"Erro: {str(e)}"}), 500


@app.route("/buscar-enviar", methods=["POST"])
def buscar_enviar():
    serial = request.form.get("serial", "").strip()
    para = request.form.get("para", "").strip()

    if not serial:
        return jsonify({"erro": "Informe o numero de serie"}), 400
    if not para:
        config = carregar_config()
        para = config.get("email_to", "")
    if not para:
        return jsonify({"erro": "Informe o email do destinatario"}), 400

    conf = carregar_config()
    if not conf.get("smtp_host") or not conf.get("email_from") or not conf.get("email_pass"):
        return jsonify({"erro": "Configure os dados SMTP em Configurar Email primeiro"}), 400

    scraper = obter_scraper()
    try:
        scraper.login()
        if serial.isdigit():
            dados = scraper.extrair_tudo(serial)
        else:
            dados = scraper.processar_serial(serial)
    finally:
        scraper._fechar()

    if not dados:
        return jsonify({"erro": "Dispositivo nao encontrado ou modelo fora do padrao"}), 404

    salvar_no_banco(dados)
    if os.path.exists(ARQUIVO_DADOS):
        existentes = carregar_json(ARQUIVO_DADOS)
        if not isinstance(existentes, list):
            existentes = [existentes]
        existentes.append(dados)
        salvar_json(existentes, ARQUIVO_DADOS)
    else:
        salvar_json([dados], ARQUIVO_DADOS)

    data_str = datetime.now().strftime("%d%m%Y_%H%M%S")
    caminho_excel = f"relatorio_contadores_{data_str}.xlsx"
    gerar_excel([dados], caminho_excel)

    try:
        ok = enviar_email(para, f"Relatorio - {serial}",
                          conf.get("mensagem", MENSAGEM_PADRAO), caminho_excel, conf)
        if ok:
            return jsonify({"ok": True, "mensagem": f"Dados de {serial} enviados para {para}!"})
        return jsonify({"erro": "Falha ao enviar email"}), 500
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/varrer-todos", methods=["POST"])
def varrer_todos():
    import threading
    from varredura_completa import varrer_todos as executar_varredura

    def tarefa():
        resultado = executar_varredura(headless=True)
        if resultado.get("aceitos", 0) > 0:
            print(f"Varredura concluida: {resultado['aceitos']} dispositivos salvos")

    thread = threading.Thread(target=tarefa, daemon=True)
    thread.start()

    return jsonify({"ok": True, "mensagem": "Varredura iniciada em segundo plano. Acompanhe o terminal."})


@app.route("/os")
def pagina_os():
    dados = carregar_os()
    resumo = obter_resumo_os()
    return render_template("os.html", dados=dados, resumo=resumo)


@app.route("/troca")
def pagina_troca():
    dados = carregar_troca()
    resumo = obter_resumo_troca()
    return render_template("troca.html", dados=dados, resumo=resumo)


@app.route("/api/os")
def api_os():
    dados = carregar_os()
    status = request.args.get("status", "")
    modelo = request.args.get("modelo", "")
    busca = request.args.get("busca", "").lower()
    if status:
        dados = [d for d in dados if d.get("status", "").lower() == status.lower()]
    if modelo:
        dados = [d for d in dados if modelo.lower() in d.get("modelo", "").lower()]
    if busca:
        dados = [d for d in dados if busca in d.get("serial", "").lower() or busca in d.get("localidade", "").lower() or busca in d.get("numero_os", "").lower()]
    return jsonify(dados)


@app.route("/api/troca")
def api_troca():
    dados = carregar_troca()
    uf = request.args.get("uf", "")
    modelo = request.args.get("modelo", "")
    busca = request.args.get("busca", "").lower()
    if uf:
        dados = [d for d in dados if d.get("uf", "").lower() == uf.lower()]
    if modelo:
        dados = [d for d in dados if modelo.lower() in d.get("modelo", "").lower()]
    if busca:
        dados = [d for d in dados if busca in d.get("serial", "").lower() or busca in d.get("loja", "").lower()]
    return jsonify(dados)


@app.route("/importar-dados", methods=["POST"])
def importar_dados():
    from importar_planilhas import importar_os, importar_troca
    try:
        qtd_os = len(importar_os())
        qtd_troca = len(importar_troca())
        return jsonify({"ok": True, "mensagem": f"Importado: {qtd_os} OS e {qtd_troca} trocas"})
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/detalhe-serial/<serial>")
def detalhe_serial(serial):
    serial = serial.upper()
    os_list = [d for d in carregar_os() if d.get("serial", "").upper() == serial]
    troca_list = [d for d in carregar_troca() if d.get("serial", "").upper() == serial]
    observacoes = obter_observacoes(serial)

    disp_info = {}
    if os.path.exists(ARQUIVO_DADOS):
        dados = carregar_json(ARQUIVO_DADOS)
        if not isinstance(dados, list):
            dados = [dados]
        for d in dados:
            disp = d.get("dispositivo", {})
            if disp.get("N\u00famero de s\u00e9rie", "").upper() == serial:
                contagens_raw = d.get("contagens", {})
                contagens = (contagens_raw.get("contagens", {})
                             if isinstance(contagens_raw, dict) else {})
                if not contagens:
                    contagens = disp.get("contagens", {})
                disp_info = {
                    "modelo": disp.get("modelo", ""),
                    "cliente": disp.get("breadcrumbs", [{}])[0].get("nome", "") if disp.get("breadcrumbs") else "",
                    "localizacao": disp.get("Localiza\u00e7\u00e3o", ""),
                    "ip": disp.get("Endere\u00e7o IP", ""),
                    "zona": disp.get("Zona", ""),
                    "ultima_atualizacao": disp.get("\u00daltima atualiza\u00e7\u00e3o", ""),
                    "contagens": contagens,
                    "consumiveis": d.get("consumiveis", {}).get("consumiveis", []),
                    "alertas_freq": d.get("alertas", {}).get("frequencia_alertas", []),
                    "alertas_anteriores": d.get("alertas", {}).get("alertas_anteriores", []),
                }
                break

    return render_template("detalhe_serial.html", serial=serial,
                           os_list=os_list, troca_list=troca_list,
                           observacoes=observacoes, disp_info=disp_info)


@app.route("/api/observacao", methods=["POST"])
def api_observacao():
    data = request.json
    serial = data.get("serial", "").upper()
    texto = data.get("texto", "").strip()
    if not serial or not texto:
        return jsonify({"erro": "Serial e texto sao obrigatorios"}), 400
    adicionar_observacao(serial, texto)
    return jsonify({"ok": True, "observacoes": obter_observacoes(serial)})


@app.route("/api/observacao/<serial>/<int:idx>", methods=["DELETE"])
def api_excluir_observacao(serial, idx):
    from importar_planilhas import carregar_observacoes, salvar_observacoes
    serial = serial.upper()
    obs = carregar_observacoes()
    if serial in obs and 0 <= idx < len(obs[serial]):
        obs[serial].pop(idx)
        if not obs[serial]:
            del obs[serial]
        salvar_observacoes(obs)
        return jsonify({"ok": True, "observacoes": obter_observacoes(serial)})
    return jsonify({"erro": "Observacao nao encontrada"}), 404


@app.route("/relatorio-os")
def relatorio_os():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    dados = carregar_os()
    wb = Workbook()
    ws = wb.active
    ws.title = "OS Chamados"

    cab = ["N\u00ba OS", "Serial", "Modelo", "CEP", "Loja", "Abertura", "Previs\u00e3o",
           "Finalizado", "Descri\u00e7\u00e3o", "Observa\u00e7\u00e3o", "Status", "Ocorr\u00eancia"]
    estilo = Font(bold=True, color="FFFFFF", size=11)
    fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, tit in enumerate(cab, 1):
        c = ws.cell(row=1, column=col, value=tit)
        c.font = estilo; c.fill = fill; c.alignment = Alignment(horizontal='center'); c.border = borda

    for i, d in enumerate(dados, 2):
        vals = [d.get(k, "") for k in ["numero_os","serial","modelo","cep","localidade","abertura","previsao","finalizado","descricao","observacao","status","ocorrencia"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.alignment = Alignment(horizontal='center'); c.border = borda

    for col, w in enumerate([14,18,16,12,30,16,16,16,30,30,14,30], 1):
        ws.column_dimensions[chr(64+col)].width = w

    caminho = "relatorio_os.xlsx"
    wb.save(caminho)
    return send_file(caminho, as_attachment=True, download_name="relatorio_os.xlsx")


@app.route("/relatorio-troca")
def relatorio_troca():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    dados = carregar_troca()
    wb = Workbook()
    ws = wb.active
    ws.title = "Troca Antecipada"

    cab = ["Loja", "CEP", "UF", "Modelo", "Serial", "Data Leitura", "% Toner",
           "Cor Toner", "IP", "Data Atend.", "Descri\u00e7\u00e3o", "Status",
           "Data Finaliza\u00e7\u00e3o", "Email", "Contato", "Retorno/A\u00e7\u00e3o", "Retorno"]
    estilo = Font(bold=True, color="FFFFFF", size=11)
    fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, tit in enumerate(cab, 1):
        c = ws.cell(row=1, column=col, value=tit)
        c.font = estilo; c.fill = fill; c.alignment = Alignment(horizontal='center'); c.border = borda

    for i, d in enumerate(dados, 2):
        vals = [d.get(k, "") for k in ["loja","cep","uf","modelo","serial","data_leitura","percentual_toner",
                                         "cor_toner","ip","data_atendimento","descricao","status",
                                         "data_finalizacao","email","contato","retorno_acao","retorno"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.alignment = Alignment(horizontal='center'); c.border = borda

    for col, w in enumerate([30,12,6,16,18,16,10,12,14,16,30,14,16,20,30,30,30], 1):
        ws.column_dimensions[chr(64+col)].width = w

    caminho = "relatorio_troca.xlsx"
    wb.save(caminho)
    return send_file(caminho, as_attachment=True, download_name="relatorio_troca.xlsx")


@app.route("/exportar-contadores")
def exportar_contadores():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    dados = []
    if os.path.exists(ARQUIVO_DADOS):
        raw = carregar_json(ARQUIVO_DADOS)
        if not isinstance(raw, list):
            raw = [raw]
        for d in raw:
            disp = d.get("dispositivo", {})
            contagens_raw = d.get("contagens", {})
            contagens = (contagens_raw.get("contagens", {})
                         if isinstance(contagens_raw, dict) else {})
            if not contagens:
                contagens = disp.get("contagens", {})
            serial = disp.get("N\u00famero de s\u00e9rie", "")
            local = disp.get("Localiza\u00e7\u00e3o", "")
            zona = disp.get("Zona", "")
            pb = contagens.get("P\u00e1ginas monocrom\u00e1ticas", "").replace(".", "")
            color = contagens.get("Colorido (equivalente A4)", "").replace(".", "")
            ult = disp.get("\u00daltima atualiza\u00e7\u00e3o", "")
            cliente = disp.get("breadcrumbs", [{}])[0].get("nome", "") if disp.get("breadcrumbs") else ""

            local_up = local.upper()
            if "FRENTE" in local_up or "CAIXA" in local_up:
                setor = "Frente de Caixa"
            elif "RM" in local_up:
                setor = "RM"
            else:
                setor = "Outro"

            dados.append((serial, pb, color, setor, local, zona, cliente, ult))

    wb = Workbook()
    ws = wb.active
    ws.title = "Contadores"

    cab = ["N\u00famero de s\u00e9rie", "Contador PB", "Contador Color",
           "Setor", "Localiza\u00e7\u00e3o", "Zona", "Cliente", "Data Atualiza\u00e7\u00e3o"]
    estilo = Font(bold=True, color="FFFFFF", size=11)
    fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, tit in enumerate(cab, 1):
        c = ws.cell(row=1, column=col, value=tit)
        c.font = estilo; c.fill = fill; c.alignment = Alignment(horizontal='center'); c.border = borda

    for i, (serial, pb, color, setor, local, zona, cliente, ult) in enumerate(dados, 2):
        vals = [serial, pb, color, setor, local, zona, cliente, ult[:16]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.alignment = Alignment(horizontal='center'); c.border = borda

    for col, w in enumerate([20, 14, 14, 16, 35, 25, 25, 18], 1):
        ws.column_dimensions[chr(64+col)].width = w

    caminho = "relatorio_contadores.xlsx"
    wb.save(caminho)
    return send_file(caminho, as_attachment=True, download_name="relatorio_contadores.xlsx")


@app.route("/relatorio-serial/<serial>")
def relatorio_serial(serial):
    serial = serial.upper()
    from relatorios import gerar_excel
    dados_os = [d for d in carregar_os() if d.get("serial", "").upper() == serial]
    dados_troca = [d for d in carregar_troca() if d.get("serial", "").upper() == serial]
    observacoes = obter_observacoes(serial)

    dados_excel = []

    if os.path.exists(ARQUIVO_DADOS):
        dados_sds = carregar_json(ARQUIVO_DADOS)
        if not isinstance(dados_sds, list):
            dados_sds = [dados_sds]
        for d in dados_sds:
            disp = d.get("dispositivo", {})
            if disp.get("N\u00famero de s\u00e9rie", "").upper() == serial:
                contagens_raw = d.get("contagens", {})
                contagens = (contagens_raw.get("contagens", {})
                             if isinstance(contagens_raw, dict) else {})
                if not contagens:
                    contagens = disp.get("contagens", {})
                dados_excel.append({
                    "dispositivo": {
                        "N\u00famero de s\u00e9rie": serial,
                        "modelo": disp.get("modelo", ""),
                        "breadcrumbs": [{"nome": disp.get("Localiza\u00e7\u00e3o", "")}],
                        "Localiza\u00e7\u00e3o": disp.get("Localiza\u00e7\u00e3o", ""),
                        "\u00daltima atualiza\u00e7\u00e3o": disp.get("\u00daltima atualiza\u00e7\u00e3o", ""),
                        "contagens": contagens
                    }
                })
                break
    for item in dados_os:
        dados_excel.append({
            "dispositivo": {
                "N\u00famero de s\u00e9rie": serial,
                "modelo": "OS",
                "breadcrumbs": [{"nome": item.get("localidade", "")}],
                "Localiza\u00e7\u00e3o": item.get("localidade", ""),
                "\u00daltima atualiza\u00e7\u00e3o": item.get("abertura", ""),
                "contagens": {
                    "N\u00ba OS": item.get("numero_os", ""),
                    "Descri\u00e7\u00e3o": item.get("descricao", ""),
                    "Status": item.get("status", ""),
                    "Observa\u00e7\u00e3o": item.get("observacao", ""),
                }
            }
        })
    for item in dados_troca:
        dados_excel.append({
            "dispositivo": {
                "N\u00famero de s\u00e9rie": serial,
                "modelo": "Troca",
                "breadcrumbs": [{"nome": item.get("loja", "")}],
                "Localiza\u00e7\u00e3o": item.get("loja", ""),
                "\u00daltima atualiza\u00e7\u00e3o": item.get("data_atendimento", ""),
                "contagens": {
                }
            }
        })

    obs_texto = "\n".join([f"[{o['data']}] {o['texto']}" for o in observacoes])
    if obs_texto:
        dados_excel.append({
            "dispositivo": {
                "N\u00famero de s\u00e9rie": serial,
                "modelo": "Observa\u00e7\u00f5es",
                "breadcrumbs": [{"nome": ""}],
                "Localiza\u00e7\u00e3o": "",
                "\u00daltima atualiza\u00e7\u00e3o": "",
                "contagens": {"Observa\u00e7\u00f5es": obs_texto}
            }
        })

    caminho = f"relatorio_{serial}.xlsx"
    gerar_excel(dados_excel, caminho)
    return send_file(caminho, as_attachment=True, download_name=f"relatorio_{serial}.xlsx")


@app.route("/api/salvar-os", methods=["POST"])
def salvar_os():
    data = request.json
    from importar_planilhas import carregar_os
    lista = carregar_os()
    campos = ["numero_os","serial","modelo","cep","localidade","abertura","previsao","finalizado","descricao","observacao","status","ocorrencia"]
    encontrado = False
    for i, item in enumerate(lista):
        if item.get("numero_os") == data.get("numero_os"):
            for k in campos:
                if k in data:
                    lista[i][k] = str(data[k])
            encontrado = True
            break
    if not encontrado:
        novo = {k: str(data.get(k, "")) for k in campos}
        lista.append(novo)
    from importar_planilhas import ARQUIVO_OS
    import json
    with open(ARQUIVO_OS, "w", encoding="utf-8") as f:
        json.dump(lista, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True, "novo": not encontrado})


@app.route("/api/salvar-troca", methods=["POST"])
def salvar_troca():
    data = request.json
    from importar_planilhas import carregar_troca
    lista = carregar_troca()
    campos = ["loja","cep","uf","modelo","serial","data_leitura","percentual_toner","cor_toner","ip","data_atendimento","descricao","status","data_finalizacao","email","contato","retorno_acao","retorno"]
    encontrado = False
    for i, item in enumerate(lista):
        if item.get("serial") == data.get("serial") and item.get("data_atendimento") == data.get("data_atendimento"):
            for k in campos:
                if k in data:
                    lista[i][k] = str(data[k])
            encontrado = True
            break
    if not encontrado:
        novo = {k: str(data.get(k, "")) for k in campos}
        lista.append(novo)
    from importar_planilhas import ARQUIVO_TROCA
    import json
    with open(ARQUIVO_TROCA, "w", encoding="utf-8") as f:
        json.dump(lista, f, ensure_ascii=False, indent=2)
    return jsonify({"ok": True, "novo": not encontrado})


if __name__ == "__main__":
    app.run(debug=False, port=5000, threaded=True)
