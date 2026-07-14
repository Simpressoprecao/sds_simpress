import os
import smtplib
import json
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

ARQUIVO_CONFIG = "config_email.json"
ARQUIVO_DADOS = "dados_dispositivo.json"

def carregar_config():
    if os.path.exists(ARQUIVO_CONFIG):
        with open(ARQUIVO_CONFIG, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def salvar_config(config):
    with open(ARQUIVO_CONFIG, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

MENSAGEM_PADRAO = """Prezados,

Segue em anexo a planilha com os dados de leitura dos dispositivos.

Atenciosamente,
Equipe de Monitoramento"""

def gerar_excel(dados_lista, caminho="relatorio_contadores.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contadores"

    cabecalho = [
        "Numero Serie", "Mes", "Ano",
        "Contador PB", "Contador Color", "Data Leitura"
    ]

    estilo_cabecalho = Font(bold=True, color="FFFFFF", size=11)
    fill_cabecalho = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, titulo in enumerate(cabecalho, 1):
        celula = ws.cell(row=1, column=col, value=titulo)
        celula.font = estilo_cabecalho
        celula.fill = fill_cabecalho
        celula.alignment = Alignment(horizontal='center')
        celula.border = borda

    linha = 2
    for dados in dados_lista:
        disp = dados.get("dispositivo", {})
        contagens = disp.get("contagens", {})
        data_leitura = disp.get("\u00daltima atualiza\u00e7\u00e3o", "")

        mes = ano = ""
        if data_leitura:
            for fmt in ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"]:
                try:
                    dt = datetime.strptime(data_leitura, fmt)
                    mes = str(dt.month)
                    ano = str(dt.year)
                    break
                except ValueError:
                    continue

        contador_pb = contagens.get("P\u00e1ginas monocrom\u00e1ticas",
                       contagens.get("Monocrom\u00e1tico (equivalente A4)", ""))
        contador_color = contagens.get("Colorido (equivalente A4)", "")

        valores = [
            disp.get("N\u00famero de s\u00e9rie", ""),
            mes,
            ano,
            contador_pb,
            contador_color,
            data_leitura
        ]

        for col, valor in enumerate(valores, 1):
            celula = ws.cell(row=linha, column=col, value=valor)
            celula.alignment = Alignment(horizontal='center')
            celula.border = borda
        linha += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 22

    wb.save(caminho)
    print(f"Planilha salva: {caminho}")
    return caminho


def enviar_email(para, assunto="Relatorio de Contadores - Dispositivos",
                 mensagem=None, anexo=None, config=None):
    if config is None:
        config = carregar_config()

    smtp_host = config.get("smtp_host", "")
    smtp_port = config.get("smtp_port", 587)
    email_from = config.get("email_from", "")
    email_pass = config.get("email_pass", "")

    if not all([smtp_host, email_from, email_pass]):
        print("Configuracao de email incompleta.")
        return False

    msg = MIMEMultipart()
    msg["From"] = email_from
    msg["To"] = para
    msg["Subject"] = assunto

    corpo = mensagem or MENSAGEM_PADRAO
    msg.attach(MIMEText(corpo, "plain", "utf-8"))

    if anexo and os.path.exists(anexo):
        with open(anexo, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename={os.path.basename(anexo)}"
            )
            msg.attach(part)

    try:
        server = smtplib.SMTP(smtp_host, smtp_port)
        server.starttls()
        server.login(email_from, email_pass)
        server.send_message(msg)
        server.quit()
        print(f"Email enviado para {para}")
        return True
    except Exception as e:
        print(f"Erro ao enviar email: {e}")
        return False


def gerar_relatorio_e_enviar(para, dados_lista=None):
    if dados_lista is None and os.path.exists(ARQUIVO_DADOS):
        with open(ARQUIVO_DADOS, "r", encoding="utf-8") as f:
            dados_lista = json.load(f)
        if not isinstance(dados_lista, list):
            dados_lista = [dados_lista]

    if not dados_lista:
        print("Nenhum dado disponivel.")
        return False

    data_str = datetime.now().strftime("%d%m%Y")
    caminho_excel = f"relatorio_contadores_{data_str}.xlsx"
    gerar_excel(dados_lista, caminho_excel)

    config = carregar_config()
    if config.get("email_to"):
        return enviar_email(
            para=para or config["email_to"],
            anexo=caminho_excel,
            config=config
        )
    print("Destinatario nao configurado.")
    return False
